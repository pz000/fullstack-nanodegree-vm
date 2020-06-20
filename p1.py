from openpyxl import load_workbook
import sqlite3
import json

def wczytajZoska(nr_miesiaca):
    plik = load_workbook(f"Zoska_{nr_miesiaca}.xlsx")
    zgloszenia = plik['Zgłoszenia']
    listaUbezpieczonych = []
    i=4
    while True:
        ubezpieczony = []
        ubezpieczony.append(zgloszenia[f'A{i}'].value)
        ubezpieczony.append(zgloszenia[f'B{i}'].value)
        ubezpieczony.append(zgloszenia[f'C{i}'].value)
        ubezpieczony.append(zgloszenia[f'D{i}'].value)
        ubezpieczony.append(zgloszenia[f'E{i}'].value)
        ubezpieczony.append(zgloszenia[f'F{i}'].value)
        ubezpieczony.append(zgloszenia[f'G{i}'].value)
        ubezpieczony.append(zgloszenia[f'H{i}'].value)
        ubezpieczony.append(zgloszenia[f'I{i}'].value)
        ubezpieczony.append(zgloszenia[f'J{i}'].value)
        ubezpieczony.append(zgloszenia[f'K{i}'].value)
        ubezpieczony.append(zgloszenia[f'L{i}'].value)
        ubezpieczony.append(zgloszenia[f'M{i}'].value)
        ubezpieczony.append(zgloszenia[f'N{i}'].value)
        ubezpieczony.append(zgloszenia[f'O{i}'].value)
        ubezpieczony.append(zgloszenia[f'P{i}'].value)
        ubezpieczony.append(zgloszenia[f'Q{i}'].value)
        listaUbezpieczonych.append(ubezpieczony)
        i+=1
        if all(x is None for x in ubezpieczony):
            break
    del listaUbezpieczonych[-1]
    for u in listaUbezpieczonych:
        for e in u:
            if u is None:
                print(u)
                print("Braki danych")
                return None
    return listaUbezpieczonych


def wczytajZoskaRezygnacje(nr_miesiaca):
    plik = load_workbook(f"Zoska_{nr_miesiaca}.xlsx")
    rezygnacje = plik['Rezygnacje']
    listaRezygnacji = []
    i=7
    while True:
        rezygnacja = []
        rezygnacja.append(rezygnacje[f'A{i}'].value)
        rezygnacja.append(rezygnacje[f'B{i}'].value)
        rezygnacja.append(rezygnacje[f'C{i}'].value)
        rezygnacja.append(rezygnacje[f'D{i}'].value)
        rezygnacja.append(rezygnacje[f'E{i}'].value)
        rezygnacja.append(rezygnacje[f'F{i}'].value)
        listaRezygnacji.append(rezygnacja)
        i+=1
        if all(x is None for x in rezygnacja):
            break
    del listaRezygnacji[-1]
    for u in listaRezygnacji:
        for e in u:
            if u is None:
                print(u)
                print("Braki danych")
                return None
    return listaRezygnacji

def stworzBazeDanych():
    conn = sqlite3.connect("myDb.db")
    c=conn.cursor()
    c.execute('drop table if exists ubezpieczeni')
    c.execute('create table if not exists ubezpieczeni(NrDeklaracji INTEGER PRIMARY KEY AUTOINCREMENT, Pesel TEXT, NrMiesiaca integer)')
    c.execute('drop table if exists rezygnacje')
    c.execute('create table if not exists rezygnacje(NrDeklaracji INTEGER, Pesel TEXT)')
    c.execute('drop table if exists wplaty')
    c.execute('create table if not exists wplaty(NrDeklaracji INTEGER, Kwota REAL, Miesiac integer)')
    conn.close()

def zarejestruj(listaUbezpieczonych, nrMiesiaca):
    conn = sqlite3.connect("myDb.db")
    c = conn.cursor()
    for u in listaUbezpieczonych:
        c.execute(f'insert into ubezpieczeni(Pesel, NrMiesiaca) values({str(u[5])}, {nrMiesiaca})')
    conn.commit()
    conn.close()

def wyrejestruj(listaRezygnacji):
    conn = sqlite3.connect("myDb.db")
    c = conn.cursor()
    for r in listaRezygnacji:
        c.execute(f'insert into rezygnacje(NrDeklaracji, Pesel) values({str(r[4])}, {str(r[4])})')
        c.execute(f'delete from ubezpieczeni where pesel = {str(r[4])}')
    conn.commit()
    conn.close()
    wyrejestrowani = []
    with open("aktualnieZarejestrowani.txt", "w") as plik:
        for w in listaRezygnacji:
            plik.write(str(w[5]))
            plik.write('\n')

def wprowadzWplaty(nrMiesiaca):
    nazwaPliku="wplaty_"+str(nrMiesiaca)+".json"
    with open(nazwaPliku) as f:
        data = json.load(f)
    conn = sqlite3.connect("myDb.db")
    c = conn.cursor()
    for e in data:
        nrPolisy=e['nrPolisy']
        kwota = e['kwota']
        naMiesiac = e['naMiesiac']
        c.execute(f'insert into wplaty(NrDeklaracji, Kwota, Miesiac) values({nrPolisy},{kwota},{naMiesiac})')
    conn.commit()
    conn.close()


def sprawdzWplaty(nrMiesiaca):
    conn = sqlite3.connect("myDb.db")
    c = conn.cursor()
    c.execute(f'delete from ubezpieczeni where NrDeklaracji not in (select NrDeklaracji from wplaty where Kwota >=70)')
    c.execute(f'SELECT SUM(Kwota) FROM wplaty where miesiac = {nrMiesiaca}')
    sumaWplat = c.fetchall()
    c.execute(f'select count() from ubezpieczeni where NrMiesiaca = {nrMiesiaca}')
    iluUbezpieczonych = c.fetchall()
    conn.close()

def czyUbezpieczony(pesel, nrMiesiaca):
    conn = sqlite3.connect("myDb.db")
    c = conn.cursor()
    c.execute(f'select * from ubezpieczeni where Pesel = {pesel} and NrMiesiaca = {nrMiesiaca}')
    czyUbezpieczony = c.fetchall()
    conn.close()
    if(len(czyUbezpieczony)==0):
        return "Nie"
    else:
        return"Tak"



stworzBazeDanych()
listaUbezpieczonych=wczytajZoska(1)
zarejestruj(listaUbezpieczonych,1)
listaRezygnacji = wczytajZoskaRezygnacje(1)
wyrejestruj(listaRezygnacji)
wprowadzWplaty(1)
sprawdzWplaty(1)
print(czyUbezpieczony(11111111111, 1))






